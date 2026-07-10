import openpyxl

from app.services import excel_exporter


def test_build_excel_happy_path_creates_file(tmp_path):
    rows = [
        {
            "tipo_busqueda": "licitaciones",
            "favorito": "SAP > 1M",
            "titulo": "Servicio de mantenimiento SAP",
            "importe": "137.400,82€",
            "detail_url": "licitaciones-ficha.php?id=1",
            "estado_extraccion": "ok",
            "error_extraccion": "",
        }
    ]
    path = excel_exporter.build_excel(rows, "licitaciones", "SAP > 1M", output_dir=tmp_path)
    assert path.exists()
    assert path.suffix == ".xlsx"
    assert path.parent == tmp_path


def test_build_excel_one_row_per_result(tmp_path):
    rows = [
        {"titulo": "Fila 1", "estado_extraccion": "ok", "error_extraccion": ""},
        {"titulo": "Fila 2", "estado_extraccion": "ok", "error_extraccion": ""},
        {"titulo": "Fila 3", "estado_extraccion": "error", "error_extraccion": "timeout"},
    ]
    path = excel_exporter.build_excel(rows, "licitaciones", "Filtro Test", output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Resultados"]
    # +1 por la fila de cabecera
    assert ws.max_row == len(rows) + 1


def test_build_excel_missing_fields_leave_cells_empty(tmp_path):
    rows = [{"titulo": "Solo tengo título"}]
    path = excel_exporter.build_excel(rows, "vencimientos", "Otro filtro", output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Resultados"]
    header = [cell.value for cell in ws[1]]
    row_values = dict(zip(header, [cell.value for cell in ws[2]]))
    assert row_values["Título"] == "Solo tengo título"
    assert row_values["Número de expediente"] in (None, "")


def test_build_excel_partial_errors_are_preserved_in_their_row(tmp_path):
    rows = [
        {"titulo": "Sin error", "estado_extraccion": "ok", "error_extraccion": ""},
        {"titulo": "Con error", "estado_extraccion": "error", "error_extraccion": "timeout cargando ficha"},
    ]
    path = excel_exporter.build_excel(rows, "licitaciones", "Filtro", output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Resultados"]
    header = [cell.value for cell in ws[1]]
    error_col = header.index("Mensaje de error del registro") + 1
    errors = [ws.cell(row=r, column=error_col).value for r in range(2, ws.max_row + 1)]
    assert "" in errors or None in errors
    assert "timeout cargando ficha" in errors


def test_build_excel_header_frozen_and_autofiltered(tmp_path):
    rows = [{"titulo": "Fila"}]
    path = excel_exporter.build_excel(rows, "licitaciones", "Filtro", output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Resultados"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_build_excel_empty_results_still_creates_file_with_headers(tmp_path):
    path = excel_exporter.build_excel([], "licitaciones", "Filtro sin resultados", output_dir=tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Resultados"]
    assert ws.max_row == 1  # solo cabecera
    assert ws.cell(row=1, column=1).value == "Tipo de búsqueda"


def test_build_excel_filename_matches_expected_pattern(tmp_path):
    rows = [{"titulo": "Fila"}]
    path = excel_exporter.build_excel(rows, "vencimientos", "SAP > 1M", output_dir=tmp_path)
    assert path.name.startswith("tenderstool_vencimientos_SAP_1M_")
    assert path.name.endswith(".xlsx")
